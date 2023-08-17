
'''
import pycel
import openpyxl
# Создаем новый документ Excel
wb = pycel.Workbook()
ws = wb.new_sheet('Sheet1')

# Заполняем ячейки значениями
ws[1, 1] = 1
ws[2, 1] = 2
ws[3, 1] = 3
ws[4, 1] = 4
ws[5, 1] = 5

# Вызываем функцию SUM для ячеек A1-A5

ws[6, 1] = '=SUM(A1:A5)'



# Выводим результат

print(ws[6, 1].value)
'''
import pycel
import openpyxl
from pycel import ExcelCompiler
wb = openpyxl.Workbook()
ws = wb.active
ws['A1'] = 1
ws['A2'] = 1
ws['A3'] = '= IF(A1>A2,"GREATER","LOWER")'


#parser.formula_references(ws['A3'])



    # Compile a formula using ExcelCompiler
excel = ExcelCompiler(excel=wb)
print(excel.evaluate('A3'))
